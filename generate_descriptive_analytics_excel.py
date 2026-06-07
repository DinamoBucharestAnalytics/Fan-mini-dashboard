import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill


BASE_DIR = Path(__file__).resolve().parent
SOURCE_PATH = BASE_DIR / "data" / "fan_survey_dashboard.xlsx"
OUTPUT_PATH = BASE_DIR / "fan_survey_descriptive_analytics.xlsx"
MAIN_SHEET = "responses_with_judet"
EXPECTED_ROWS = 10499

COUNTRY_NORMALIZE = {
    "romania": "Romania",
    "r moldova": "Moldova",
    "republica moldova": "Moldova",
    "moldova": "Moldova",
    "tarile de jos": "Netherlands",
    "olanda": "Netherlands",
    "italia": "Italy",
    "spania": "Spain",
    "germania": "Germany",
    "franta": "France",
    "franţa": "France",
    "belgia": "Belgium",
    "irlanda": "Ireland",
    "marea britanie": "United Kingdom",
    "regatul unit": "United Kingdom",
    "anglia": "United Kingdom",
    "sua": "United States",
    "statele unite": "United States",
    "statele unite ale americii": "United States",
    "canada": "Canada",
    "austria": "Austria",
    "suedia": "Sweden",
    "elvetia": "Switzerland",
    "danemarca": "Denmark",
    "norvegia": "Norway",
    "portugalia": "Portugal",
    "grecia": "Greece",
    "cipru": "Cyprus",
    "turcia": "Turkey",
    "ucraina": "Ukraine",
    "cehia": "Czechia",
    "luxemburg": "Luxembourg",
    "thailanda": "Thailand",
    "malta": "Malta",
    "lituania": "Lithuania",
    "islanda": "Iceland",
    "noua zeelanda": "New Zealand",
    "polonia": "Poland",
    "filipine": "Philippines",
    "finlanda": "Finland",
    "japonia": "Japan",
    "japan": "Japan",
    "reunion": "Réunion",
    "réunion": "Réunion",
    "serbia": "Serbia",
    "gibraltar": "Gibraltar",
    "emiratele arabe unite": "United Arab Emirates",
    "eau": "United Arab Emirates",
    "united arab emirates": "United Arab Emirates",
    "australia": "Australia",
    "afganistan": "Afghanistan",
    "afghanistan": "Afghanistan",
}


def normalize_text(value: object) -> str:
    if pd.isna(value):
        return ""
    text = str(value)
    text = text.replace("ş", "s").replace("ţ", "t").replace("ș", "s").replace("ț", "t")
    text = "".join(
        char
        for char in unicodedata.normalize("NFKD", text)
        if not unicodedata.combining(char)
    )
    text = re.sub(r"[^A-Za-z0-9\s]", " ", text).lower().strip()
    return re.sub(r"\s+", " ", text)


def clean_values(series: pd.Series) -> pd.Series:
    values = series.dropna().astype(str).str.strip()
    return values[values.ne("") & values.str.lower().ne("nan")]


def count_table(
    df: pd.DataFrame,
    col: str,
    label_col: str | None = None,
    percentage_col: str = "percentage",
) -> pd.DataFrame:
    values = clean_values(df[col])
    label = label_col or col
    out = values.value_counts().reset_index()
    out.columns = [label, "count"]
    denominator = int(out["count"].sum())
    out[percentage_col] = out["count"] / denominator if denominator else 0
    return out


def geographic_county(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    out = count_table(
        df,
        "Județ atribuit",
        percentage_col="% of valid county responses",
    )
    out["% of all respondents"] = out["count"] / total if total else 0
    return out


def geographic_country(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    raw = clean_values(df["Țară de reședință"]).rename("Țară de reședință")
    work = raw.to_frame()
    work["country_normalized"] = work["Țară de reședință"].map(
        lambda value: COUNTRY_NORMALIZE.get(normalize_text(value), value)
    )
    out = (
        work.groupby(["Țară de reședință", "country_normalized"], dropna=False)
        .size()
        .reset_index(name="count")
        .sort_values("count", ascending=False)
        .reset_index(drop=True)
    )
    valid = int(out["count"].sum())
    out["% of valid country responses"] = out["count"] / valid if valid else 0
    out["% of all respondents"] = out["count"] / total if total else 0
    return out


def one_word_phrases(df: pd.DataFrame) -> pd.DataFrame:
    return count_table(
        df,
        "Dinamo un cuvânt - normalizat",
        percentage_col="percentage",
    )


def one_word_categories(df: pd.DataFrame) -> pd.DataFrame:
    return count_table(
        df,
        "Dinamo un cuvânt - categorie",
        percentage_col="percentage",
    )


def logo_mentioned(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    mentioned = int(df["Siglă menționată"].fillna("").astype(str).str.strip().eq("Da").sum())
    out = pd.DataFrame(
        {
            "Siglă menționată": ["Da", "Nu"],
            "count": [mentioned, total - mentioned],
        }
    )
    out["% of all respondents"] = out["count"] / total if total else 0
    return out


def logo_sentiment(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    order = ["sigla_pozitiv", "sigla_negativ", "sigla_mixt"]
    counts = clean_values(df["Siglă sentiment"]).value_counts()
    out = pd.DataFrame(
        {
            "Siglă sentiment": order,
            "count": [int(counts.get(item, 0)) for item in order],
        }
    )
    out["% of all respondents"] = out["count"] / total if total else 0
    return out


def logo_source_columns(df: pd.DataFrame) -> pd.DataFrame:
    sources: list[str] = []
    for value in clean_values(df["Coloană mențiune siglă"]):
        for source in value.split(";"):
            source = source.strip()
            if source:
                sources.append(source)
    mentioned = int(df["Siglă menționată"].fillna("").astype(str).str.strip().eq("Da").sum())
    if not sources:
        return pd.DataFrame(
            columns=[
                "Source column",
                "count",
                "% of logo source references",
                "% of logo-mentioned respondents",
            ]
        )
    out = pd.Series(sources).value_counts().reset_index()
    out.columns = ["Source column", "count"]
    source_total = int(out["count"].sum())
    out["% of logo source references"] = out["count"] / source_total if source_total else 0
    out["% of logo-mentioned respondents"] = out["count"] / mentioned if mentioned else 0
    return out


def summary_table(df: pd.DataFrame) -> pd.DataFrame:
    total = len(df)
    logo_mentions = int(df["Siglă menționată"].fillna("").astype(str).str.strip().eq("Da").sum())
    metrics = [
        ("total respondents", total),
        ("valid county responses", int(clean_values(df["Județ atribuit"]).shape[0])),
        ("valid country responses", int(clean_values(df["Țară de reședință"]).shape[0])),
        ("valid one-word responses", int(clean_values(df["Dinamo un cuvânt - normalizat"]).shape[0])),
        ("total logo mentions", logo_mentions),
    ]
    return pd.DataFrame(metrics, columns=["metric", "value"])


def write_workbook(sheets: dict[str, pd.DataFrame], output_path: Path) -> None:
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, data in sheets.items():
            data.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill("solid", fgColor="E30613")
    header_font = Font(color="FFFFFF", bold=True)
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in ws.columns:
            header = str(column_cells[0].value)
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_len + 2, len(header) + 2), 60)
            if "%" in header or header == "percentage":
                for cell in column_cells[1:]:
                    cell.number_format = "0.0%"
    wb.save(output_path)


def validate(df: pd.DataFrame, sheets: dict[str, pd.DataFrame]) -> None:
    if len(df) != EXPECTED_ROWS:
        raise ValueError(f"Expected {EXPECTED_ROWS} source rows, found {len(df)}")
    checks = {
        "county": (int(sheets["geographic_county"]["count"].sum()), int(clean_values(df["Județ atribuit"]).shape[0])),
        "country": (int(sheets["geographic_country"]["count"].sum()), int(clean_values(df["Țară de reședință"]).shape[0])),
        "one_word_phrases": (
            int(sheets["one_word_phrases"]["count"].sum()),
            int(clean_values(df["Dinamo un cuvânt - normalizat"]).shape[0]),
        ),
        "one_word_categories": (
            int(sheets["one_word_categories"]["count"].sum()),
            int(clean_values(df["Dinamo un cuvânt - categorie"]).shape[0]),
        ),
        "logo_mentioned": (int(sheets["logo_mentioned"]["count"].sum()), len(df)),
    }
    for name, (actual, expected) in checks.items():
        if actual != expected:
            raise ValueError(f"{name} count mismatch: {actual} != {expected}")


def main() -> None:
    df = pd.read_excel(SOURCE_PATH, sheet_name=MAIN_SHEET)
    sheets = {
        "summary": summary_table(df),
        "geographic_county": geographic_county(df),
        "geographic_country": geographic_country(df),
        "one_word_phrases": one_word_phrases(df),
        "one_word_categories": one_word_categories(df),
        "logo_mentioned": logo_mentioned(df),
        "logo_sentiment": logo_sentiment(df),
        "logo_source_columns": logo_source_columns(df),
    }
    validate(df, sheets)
    write_workbook(sheets, OUTPUT_PATH)
    written = pd.ExcelFile(OUTPUT_PATH).sheet_names
    expected = list(sheets)
    if written != expected:
        raise ValueError(f"Unexpected sheets: {written}")
    print(f"Wrote {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
